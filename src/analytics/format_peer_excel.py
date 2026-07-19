import os
from statistics import median
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

FILE = "output/peer_comparison.xlsx"

print("Loading workbook...")

wb = load_workbook(FILE)

# Colors
GREEN = PatternFill(fill_type="solid", start_color="90EE90")
YELLOW = PatternFill(fill_type="solid", start_color="FFF59D")
RED = PatternFill(fill_type="solid", start_color="FFB6B6")
GOLD = PatternFill(fill_type="solid", start_color="FFD700")
BLUE = PatternFill(
    fill_type="solid",
    start_color="ADD8E6"
)
for ws in wb.worksheets:

    print(f"Formatting {ws.title}")

    # Get headers
    headers = [cell.value for cell in ws[1]]

    # Benchmark column
    benchmark_col = None

    if "is_benchmark" in headers:
        benchmark_col = headers.index("is_benchmark") + 1

    # Highlight benchmark row
    if benchmark_col:

        for row in range(2, ws.max_row + 1):

            value = ws.cell(row=row, column=benchmark_col).value

            if value == 1:

                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = GOLD

    # Color percentile columns
    for col in range(1, ws.max_column + 1):

        header = ws.cell(row=1, column=col).value

        if header and str(header).endswith("_percentile"):

            for row in range(2, ws.max_row + 1):

                value = ws.cell(row=row, column=col).value

                if value is None:
                    continue

                try:

                    value = float(value)

                    if value >= 75:
                        ws.cell(row=row, column=col).fill = GREEN

                    elif value >= 25:
                        ws.cell(row=row, column=col).fill = YELLOW

                    else:
                        ws.cell(row=row, column=col).fill = RED

                except:
                    pass
    # -----------------------------------
    # Add Median Summary Row
    # -----------------------------------

    median_row = ws.max_row + 1

    ws.cell(row=median_row, column=1).value = "Median"
    for col in range(1, ws.max_column + 1):
        ws.cell(
            row=median_row,
            column=col
        ).fill = BLUE

    for col in range(2, ws.max_column + 1):

        values = []

        for row in range(2, ws.max_row):

            value = ws.cell(row=row, column=col).value

            if isinstance(value, (int, float)):
                values.append(value)

        if values:
            ws.cell(
                row=median_row,
                column=col
            ).value = round(median(values), 2)
wb.save(FILE)

print("\npeer_comparison.xlsx formatted successfully!")